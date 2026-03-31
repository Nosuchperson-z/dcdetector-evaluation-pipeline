import argparse
import csv
import json
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sklearn.metrics import f1_score
import torch

from unified_evaluator import (
    aggregate_window_point_scores_to_points,
    apply_smoothing,
    composite_event_f1,
    covered_prefix_length,
    detection_delay,
    event_detection_delays,
    event_scores,
    events_to_points,
    filter_events_by_score,
    filter_short_events,
    merge_close_events,
    normalized_delay,
    points_to_events,
    threshold_candidates,
    threshold_scores,
)


ROOT = Path(__file__).resolve().parents[2]
PROCESSED_ROOT = ROOT / "data_processed"
TABLE_DIR = ROOT / "outputs" / "tables"
OFFICIAL_TABLE_PATH = TABLE_DIR / "official_raw_metrics.csv"
KDD_ROOT = ROOT / "KDD2023-DCdetector"

if str(KDD_ROOT) not in sys.path:
    sys.path.insert(0, str(KDD_ROOT))

from solver import my_kl_loss  # noqa: E402
from model.DCdetector import DCdetector  # noqa: E402


MODEL_CONFIGS = {
    "SMAP": {
        "win_size": 105,
        "patch_size": [3, 5, 7],
        "batch_size": 8,
        "input_c": 25,
        "output_c": 25,
    },
    "MSL": {
        "win_size": 90,
        "patch_size": [3, 5],
        "batch_size": 8,
        "input_c": 55,
        "output_c": 55,
    },
    "HAI21.03": {
        "win_size": 100,
        "patch_size": [2, 4, 5, 10, 20, 25, 50],
        "batch_size": 4,
        "input_c": 57,
        "output_c": 57,
    },
}


CALIBRATION_SCORE_CACHE: dict[tuple, np.ndarray] = {}


def parse_args():
    parser = argparse.ArgumentParser(description="Run Event-aware DCdetector v2 on saved score artifacts.")
    parser.add_argument("--config", default=str(ROOT / "configs" / "eval" / "event_aware_v2.json"))
    parser.add_argument("--datasets", nargs="*", default=None)
    return parser.parse_args()


def load_config(path: str) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def load_official_metrics(path: Path | None = None) -> dict[str, dict[str, str]]:
    csv_path = OFFICIAL_TABLE_PATH if path is None else Path(path)
    with csv_path.open("r", encoding="utf-8", newline="") as fp:
        rows = list(csv.DictReader(fp))
    normalized = {}
    for row in rows:
        if "PA-F1" not in row and "F1" in row:
            row["PA-F1"] = row["F1"]
        row.setdefault("VUS-PR", "NA")
        row.setdefault("VUS-ROC", "NA")
        normalized[row["Dataset"]] = row
    return normalized


def default_official_metrics() -> dict[str, str]:
    return {
        "PA-F1": "NA",
        "VUS-PR": "NA",
        "VUS-ROC": "NA",
    }


def resolve_score_root(dataset: str, config: dict) -> Path:
    override = config.get("score_source_overrides", {}).get(dataset)
    if override:
        return ROOT / override
    return ROOT / config["score_root"]


def resolve_score_file(dataset: str, score_root: Path, filename: str) -> Path:
    dataset_dir = score_root / dataset
    direct = dataset_dir / filename
    if direct.exists():
        return direct
    matches = sorted(dataset_dir.glob(filename.replace(".npy", "*.npy")))
    if not matches:
        raise FileNotFoundError(f"Missing score file for {dataset}: {filename}")
    return matches[0]


def format_value(value, digits: int = 4):
    if isinstance(value, str):
        return value
    if value is None:
        return "NA"
    if np.isnan(value):
        return "NA"
    if np.isposinf(value):
        return "inf"
    if np.isneginf(value):
        return "-inf"
    return f"{float(value):.{digits}f}"


def resolve_device(requested_device: str = "auto") -> torch.device:
    if requested_device == "cpu":
        return torch.device("cpu")
    if requested_device == "cuda":
        return torch.device("cuda:0")
    return torch.device("cuda:0" if torch.cuda.is_available() else "cpu")


def build_model_bundle(dataset: str, requested_device: str = "auto") -> tuple[torch.nn.Module, torch.device, int, int]:
    cfg = MODEL_CONFIGS[dataset]
    device = resolve_device(requested_device)
    model = DCdetector(
        win_size=cfg["win_size"],
        enc_in=cfg["input_c"],
        c_out=cfg["output_c"],
        n_heads=1,
        d_model=256,
        e_layers=3,
        patch_size=cfg["patch_size"],
        channel=cfg["input_c"],
    ).to(device)
    checkpoint = torch.load(KDD_ROOT / "checkpoints" / f"{dataset}_checkpoint.pth", map_location=device)
    model.load_state_dict(checkpoint, strict=False)
    model.eval()
    return model, device, int(cfg["win_size"]), int(cfg["batch_size"])


def make_overlapping_windows(sequence: np.ndarray, win_size: int) -> tuple[np.ndarray, np.ndarray]:
    if sequence.shape[0] < win_size:
        raise ValueError("sequence is shorter than win_size")
    starts = np.arange(0, sequence.shape[0] - win_size + 1, dtype=np.int64)
    windows = np.lib.stride_tricks.sliding_window_view(sequence, window_shape=win_size, axis=0)
    windows = np.moveaxis(windows, -1, 1).astype(np.float32, copy=False)
    end_indices = starts + win_size - 1
    return windows, end_indices


def compute_window_point_scores(model, device, windows: np.ndarray, win_size: int, batch_size: int) -> np.ndarray:
    temperature = 50
    outputs = []
    model.eval()
    with torch.no_grad():
        for start in range(0, windows.shape[0], batch_size):
            batch = torch.from_numpy(windows[start:start + batch_size]).float().to(device)
            series, prior = model(batch)
            series_loss = 0.0
            prior_loss = 0.0
            for level in range(len(prior)):
                normalized_prior = prior[level] / torch.unsqueeze(torch.sum(prior[level], dim=-1), dim=-1).repeat(
                    1, 1, 1, win_size
                )
                if level == 0:
                    series_loss = my_kl_loss(series[level], normalized_prior.detach()) * temperature
                    prior_loss = my_kl_loss(normalized_prior, series[level].detach()) * temperature
                else:
                    series_loss += my_kl_loss(series[level], normalized_prior.detach()) * temperature
                    prior_loss += my_kl_loss(normalized_prior, series[level].detach()) * temperature
            metric = torch.softmax((-series_loss - prior_loss), dim=-1)
            outputs.append(metric.detach().cpu().numpy().astype(np.float32))
    return np.concatenate(outputs, axis=0)


def load_calibration_sequence(dataset: str, calibration_cfg: dict) -> np.ndarray:
    val_fraction = float(calibration_cfg.get("train_split_fraction", 0.1))
    if dataset in {"SMAP", "MSL"}:
        train_data = np.load(KDD_ROOT / "dataset" / dataset / f"{dataset}_train.npy").astype(np.float32)
        split_index = int(train_data.shape[0] * (1 - val_fraction))
        split_index = max(1, min(split_index, train_data.shape[0] - 1))
        return np.asarray(train_data[split_index:], dtype=np.float32)
    return np.load(PROCESSED_ROOT / dataset / "val.npy").astype(np.float32)


def load_calibration_scores_from_artifacts(
    dataset: str,
    score_root: Path | None,
    method_cfg: dict,
    calibration_cfg: dict,
) -> np.ndarray | None:
    if score_root is None:
        return None
    train_score_path = score_root / dataset / "train_point_scores.npy"
    if not train_score_path.exists():
        return None

    train_point_scores = np.load(train_score_path).astype(np.float32).reshape(-1)
    if train_point_scores.size <= 1:
        return None

    val_fraction = float(calibration_cfg.get("train_split_fraction", 0.1))
    split_index = int(train_point_scores.shape[0] * (1 - val_fraction))
    split_index = max(1, min(split_index, train_point_scores.shape[0] - 1))
    calibration_scores = np.asarray(train_point_scores[split_index:], dtype=np.float32)
    return apply_smoothing(calibration_scores, method_cfg["smoothing"], method_cfg["smoothing_param"])


def load_calibration_point_scores(
    dataset: str,
    score_root: Path | None,
    method_cfg: dict,
    calibration_cfg: dict,
    requested_device: str = "auto",
) -> np.ndarray:
    cache_key = (
        dataset,
        str(score_root) if score_root is not None else "none",
        calibration_cfg.get("mode", "prefix_fraction"),
        float(calibration_cfg.get("train_split_fraction", 0.1)),
        method_cfg["aggregation"],
        method_cfg["smoothing"],
        float(method_cfg["smoothing_param"]),
        requested_device,
    )
    cached = CALIBRATION_SCORE_CACHE.get(cache_key)
    if cached is not None:
        return cached.copy()

    artifact_scores = load_calibration_scores_from_artifacts(dataset, score_root, method_cfg, calibration_cfg)
    if artifact_scores is not None:
        CALIBRATION_SCORE_CACHE[cache_key] = artifact_scores.copy()
        return artifact_scores

    sequence = load_calibration_sequence(dataset, calibration_cfg)
    model, device, win_size, batch_size = build_model_bundle(dataset, requested_device=requested_device)
    windows, end_indices = make_overlapping_windows(sequence, win_size)
    window_point_scores = compute_window_point_scores(model, device, windows, win_size=win_size, batch_size=batch_size)
    point_scores = build_point_scores(
        labels=np.zeros(sequence.shape[0], dtype=np.int64),
        window_point_scores=window_point_scores,
        window_end_indices=end_indices,
        aggregation=method_cfg["aggregation"],
        smoothing=method_cfg["smoothing"],
        smoothing_param=method_cfg["smoothing_param"],
    )
    CALIBRATION_SCORE_CACHE[cache_key] = point_scores.copy()
    del model
    if device.type == "cuda":
        torch.cuda.empty_cache()
    return point_scores


def load_dataset_inputs(dataset: str, score_root: Path, trim_to_coverage: bool):
    labels = np.load(PROCESSED_ROOT / dataset / "label.npy").astype(np.int64)
    window_point_scores = np.load(resolve_score_file(dataset, score_root, "test_window_point_scores.npy")).astype(np.float32)
    window_end_indices = np.load(resolve_score_file(dataset, score_root, "test_window_end_indices.npy")).astype(np.int64).reshape(-1)

    effective_length = min(window_point_scores.shape[0], window_end_indices.shape[0])
    window_point_scores = window_point_scores[:effective_length]
    window_end_indices = window_end_indices[:effective_length]

    covered_length = covered_prefix_length(window_end_indices, len(labels))
    original_length = int(len(labels))
    if trim_to_coverage:
        labels = labels[:covered_length]

    return {
        "labels": labels,
        "window_point_scores": window_point_scores,
        "window_end_indices": window_end_indices,
        "covered_length": int(covered_length),
        "original_length": original_length,
    }


def calibration_prefix_length(total_length: int, calibration_cfg: dict) -> int:
    if total_length <= 0:
        return 0
    if not bool(calibration_cfg.get("enabled", False)):
        return total_length
    mode = calibration_cfg.get("mode", "prefix_fraction")
    if mode != "prefix_fraction":
        raise ValueError(f"Unsupported calibration mode: {mode}")
    fraction = float(calibration_cfg.get("fraction", 0.2))
    min_points = int(calibration_cfg.get("min_points", 1))
    max_points = calibration_cfg.get("max_points")
    calib_length = max(min_points, int(np.ceil(total_length * fraction)))
    if max_points is not None:
        calib_length = min(calib_length, int(max_points))
    return max(1, min(int(total_length), int(calib_length)))


def test_holdout_offset(total_length: int, calibration_cfg: dict) -> int:
    if total_length <= 0:
        return 0
    holdout_fraction = float(calibration_cfg.get("test_holdout_fraction", 0.2))
    min_points = int(calibration_cfg.get("test_holdout_min_points", 1))
    offset = max(min_points, int(np.ceil(total_length * holdout_fraction)))
    if offset >= total_length:
        return 0
    return int(max(0, offset))


def build_threshold_views(
    dataset: str,
    labels: np.ndarray,
    point_scores: np.ndarray,
    calibration_cfg: dict,
    method_cfg: dict,
    requested_device: str = "auto",
    score_root: Path | None = None,
) -> dict:
    mode = calibration_cfg.get("mode", "prefix_fraction")
    if mode == "train_split_normal_quantile":
        calibration_scores = load_calibration_point_scores(
            dataset=dataset,
            score_root=score_root,
            method_cfg=method_cfg,
            calibration_cfg=calibration_cfg,
            requested_device=requested_device,
        )
        evaluate_on_holdout_only = bool(calibration_cfg.get("evaluate_on_holdout_only", True))
        evaluation_offset = test_holdout_offset(len(labels), calibration_cfg) if evaluate_on_holdout_only else 0
        return {
            "selection_labels": None,
            "selection_scores": calibration_scores,
            "selection_length": int(len(calibration_scores)),
            "selection_source": "validation_normal",
            "evaluation_labels": labels[evaluation_offset:],
            "evaluation_scores": point_scores[evaluation_offset:],
            "evaluation_offset": int(evaluation_offset),
            "evaluate_on_holdout_only": bool(evaluate_on_holdout_only and evaluation_offset > 0),
        }

    total_length = int(len(labels))
    selection_length = calibration_prefix_length(total_length, calibration_cfg)
    if not bool(calibration_cfg.get("enabled", False)):
        selection_length = total_length
    evaluate_on_holdout_only = bool(calibration_cfg.get("evaluate_on_holdout_only", True))
    evaluation_offset = selection_length if evaluate_on_holdout_only and selection_length < total_length else 0
    return {
        "selection_labels": labels[:selection_length],
        "selection_scores": point_scores[:selection_length],
        "selection_length": int(selection_length),
        "selection_source": "test_prefix" if selection_length < total_length else "test_full",
        "evaluation_labels": labels[evaluation_offset:],
        "evaluation_scores": point_scores[evaluation_offset:],
        "evaluation_offset": int(evaluation_offset),
        "evaluate_on_holdout_only": evaluate_on_holdout_only and evaluation_offset > 0,
    }


def select_point_threshold(selection_scores: np.ndarray, selection_labels: np.ndarray | None, method_cfg: dict, calibration_cfg: dict) -> float:
    if selection_labels is None:
        method = str(method_cfg.get("threshold_method", "quantile"))
        param = method_cfg.get("threshold_param")
        if param is None:
            if method == "quantile":
                param = float(calibration_cfg.get("point_score_quantile", 0.995))
            elif method == "dynamic":
                param = float(calibration_cfg.get("point_score_dynamic_z", 3.0))
        return threshold_scores(selection_scores, method, labels=None, param=param)
    return threshold_scores(
        selection_scores,
        method_cfg["threshold_method"],
        labels=selection_labels,
        param=method_cfg["threshold_param"],
    )


def select_event_aware_v2_thresholds(
    selection_scores: np.ndarray,
    selection_labels: np.ndarray | None,
    cfg: dict,
    calibration_cfg: dict,
) -> tuple[dict, list[dict]]:
    if selection_labels is not None:
        return search_event_aware_v2(selection_labels, selection_scores, cfg)

    point_method = str(cfg.get("nolabel_point_threshold_method", "quantile"))
    point_param = cfg.get("nolabel_point_threshold_param")
    if point_param is None:
        if point_method == "quantile":
            point_param = float(calibration_cfg.get("point_score_quantile", 0.995))
        elif point_method == "dynamic":
            point_param = float(calibration_cfg.get("point_score_dynamic_z", 3.0))
    point_threshold = threshold_scores(selection_scores, point_method, labels=None, param=point_param)
    raw_preds = (selection_scores >= point_threshold).astype(np.int64)
    raw_events = points_to_events(raw_preds)
    merged_events = merge_close_events(raw_events, int(cfg["gap_size"])) if cfg["gap_merge"] else list(raw_events)
    length_filtered_events = filter_short_events(merged_events, int(cfg["min_event_length"]))
    pooled_scores = event_scores(selection_scores, length_filtered_events, pooling=cfg["event_score_pooling"])

    event_method = str(cfg.get("nolabel_event_threshold_method", "quantile"))
    event_param = cfg.get("nolabel_event_threshold_param")
    event_threshold = None
    if event_method != "none" and pooled_scores.size:
        if event_param is None:
            if event_method == "quantile":
                event_param = float(calibration_cfg.get("event_score_quantile", 0.9))
            elif event_method == "dynamic":
                event_param = float(calibration_cfg.get("event_score_dynamic_z", 2.0))
        event_threshold = threshold_scores(pooled_scores, event_method, labels=None, param=event_param)

    processed = postprocess_events(
        point_scores=selection_scores,
        raw_events=raw_events,
        gap_merge_enabled=bool(cfg["gap_merge"]),
        gap_size=int(cfg["gap_size"]),
        min_event_length=int(cfg["min_event_length"]),
        event_threshold=event_threshold,
        event_score_pooling=cfg["event_score_pooling"],
    )
    preds = events_to_points(processed["final_events"], len(selection_scores))
    positive_ratio = float(np.mean(preds)) if preds.size else 0.0
    event_threshold_value = float(event_threshold) if event_threshold is not None else None
    best_row = {
        "point_threshold": float(point_threshold),
        "event_threshold": event_threshold_value,
        "objective": float(1.0 - positive_ratio),
        "f1": float("nan"),
        "fc1": float("nan"),
        "delay": float("nan"),
        "preds": preds,
        "metrics": {
            "threshold": float(point_threshold),
            "f1": float("nan"),
            "fc1": float("nan"),
            "delay": float("nan"),
            "predicted_events": points_to_events(preds),
            "true_events": [],
            "normalized_delay": float("nan"),
            "aupr_proxy": float(np.mean(selection_scores[preds == 1])) if np.any(preds == 1) else 0.0,
        },
        "processed": processed,
    }
    curve_rows = [
        {
            "point_threshold": float(point_threshold),
            "f1": float("nan"),
            "fc1": float("nan"),
            "delay": float("nan"),
            "objective": float(1.0 - positive_ratio),
            "event_threshold": float(event_threshold) if event_threshold is not None else float("nan"),
        }
    ]
    return best_row, curve_rows


def build_point_scores(
    labels: np.ndarray,
    window_point_scores: np.ndarray,
    window_end_indices: np.ndarray,
    aggregation: str,
    smoothing: str,
    smoothing_param: float,
) -> np.ndarray:
    point_scores = aggregate_window_point_scores_to_points(
        window_point_scores=window_point_scores,
        window_end_indices=window_end_indices,
        total_length=len(labels),
        method=aggregation,
    )
    return apply_smoothing(point_scores, smoothing, smoothing_param)


def evaluate_predictions(labels: np.ndarray, preds: np.ndarray, point_scores: np.ndarray, threshold=None) -> dict:
    return {
        "threshold": float(threshold) if threshold is not None else float("nan"),
        "f1": float(f1_score(labels, preds, zero_division=0)),
        "fc1": float(composite_event_f1(labels, preds)),
        "delay": float(detection_delay(labels, preds)),
        "predicted_events": points_to_events(preds),
        "true_events": points_to_events(labels),
        "normalized_delay": float(normalized_delay(labels, preds)),
        "aupr_proxy": float(np.mean(point_scores[preds == 1])) if np.any(preds == 1) else 0.0,
    }


def objective_value(labels: np.ndarray, preds: np.ndarray, objective: str, objective_lambda: float) -> float:
    fc1 = composite_event_f1(labels, preds)
    delay_penalty = normalized_delay(labels, preds)
    if objective == "fc1_minus_lambda_delay":
        return float(fc1 - objective_lambda * delay_penalty)
    if objective == "fc1_then_delay":
        return float(fc1 - 1e-6 * delay_penalty)
    raise ValueError(f"Unsupported threshold objective: {objective}")


def event_threshold_candidates(values: np.ndarray, method: str, steps: int) -> np.ndarray:
    if values.size == 0:
        return np.asarray([0.0], dtype=np.float64)
    if method == "none":
        return np.asarray([float(values.min())], dtype=np.float64)
    if method == "sweep":
        quantiles = np.linspace(0.0, 1.0, steps, dtype=np.float64)
        return np.unique(np.quantile(values, quantiles)).astype(np.float64)
    raise ValueError(f"Unsupported event threshold method: {method}")


def postprocess_events(
    point_scores: np.ndarray,
    raw_events: list[tuple[int, int]],
    gap_merge_enabled: bool,
    gap_size: int,
    min_event_length: int,
    event_threshold: float | None,
    event_score_pooling: str,
) -> dict:
    merged_events = merge_close_events(raw_events, gap_size) if gap_merge_enabled else list(raw_events)
    length_filtered_events = filter_short_events(merged_events, min_event_length)
    if event_threshold is None:
        final_events = list(length_filtered_events)
    else:
        final_events = filter_events_by_score(
            point_scores=point_scores,
            events=length_filtered_events,
            threshold=float(event_threshold),
            pooling=event_score_pooling,
        )
    return {
        "raw_events": raw_events,
        "merged_events": merged_events,
        "length_filtered_events": length_filtered_events,
        "final_events": final_events,
    }


def apply_event_aware_thresholds(
    point_scores: np.ndarray,
    total_length: int,
    cfg: dict,
    point_threshold: float,
    event_threshold: float | None,
) -> dict:
    raw_preds = (point_scores >= float(point_threshold)).astype(np.int64)
    raw_events = points_to_events(raw_preds)
    processed = postprocess_events(
        point_scores=point_scores,
        raw_events=raw_events,
        gap_merge_enabled=bool(cfg["gap_merge"]),
        gap_size=int(cfg["gap_size"]),
        min_event_length=int(cfg["min_event_length"]),
        event_threshold=float(event_threshold) if event_threshold is not None else None,
        event_score_pooling=cfg["event_score_pooling"],
    )
    preds = events_to_points(processed["final_events"], total_length)
    return {
        "preds": preds,
        "processed": processed,
    }


def search_event_aware_v2(labels: np.ndarray, point_scores: np.ndarray, cfg: dict) -> tuple[dict, list[dict]]:
    point_thresholds = threshold_candidates(point_scores, steps=int(cfg["point_threshold_steps"]))
    best_row = None
    curve_rows = []

    for point_threshold in point_thresholds:
        raw_preds = (point_scores >= float(point_threshold)).astype(np.int64)
        raw_events = points_to_events(raw_preds)
        merged_events = merge_close_events(raw_events, int(cfg["gap_size"])) if cfg["gap_merge"] else list(raw_events)
        length_filtered_events = filter_short_events(merged_events, int(cfg["min_event_length"]))
        pooled_scores = event_scores(point_scores, length_filtered_events, pooling=cfg["event_score_pooling"])
        event_thresholds = event_threshold_candidates(
            pooled_scores,
            method=cfg["event_threshold_method"],
            steps=int(cfg["event_threshold_steps"]),
        )

        best_for_point = None
        for event_threshold in event_thresholds:
            if length_filtered_events:
                final_events = [
                    event for event, score in zip(length_filtered_events, pooled_scores) if float(score) >= float(event_threshold)
                ]
            else:
                final_events = []
            preds = events_to_points(final_events, len(labels))
            metrics = evaluate_predictions(labels, preds, point_scores, threshold=float(point_threshold))
            objective = objective_value(
                labels=labels,
                preds=preds,
                objective=cfg["threshold_objective"],
                objective_lambda=float(cfg["threshold_objective_lambda"]),
            )
            row = {
                "point_threshold": float(point_threshold),
                "event_threshold": float(event_threshold),
                "objective": float(objective),
                "f1": float(metrics["f1"]),
                "fc1": float(metrics["fc1"]),
                "delay": float(metrics["delay"]),
                "preds": preds,
                "metrics": metrics,
                "processed": {
                    "raw_events": raw_events,
                    "merged_events": merged_events,
                    "length_filtered_events": length_filtered_events,
                    "final_events": final_events,
                },
            }
            if best_for_point is None or (
                row["objective"],
                row["fc1"],
                row["f1"],
                -row["delay"] if np.isfinite(row["delay"]) else float("-inf"),
            ) > (
                best_for_point["objective"],
                best_for_point["fc1"],
                best_for_point["f1"],
                -best_for_point["delay"] if np.isfinite(best_for_point["delay"]) else float("-inf"),
            ):
                best_for_point = row

            if best_row is None or (
                row["objective"],
                row["fc1"],
                row["f1"],
                -row["delay"] if np.isfinite(row["delay"]) else float("-inf"),
            ) > (
                best_row["objective"],
                best_row["fc1"],
                best_row["f1"],
                -best_row["delay"] if np.isfinite(best_row["delay"]) else float("-inf"),
            ):
                best_row = row

        curve_rows.append(
            {
                "point_threshold": float(point_threshold),
                "f1": float(best_for_point["f1"]),
                "fc1": float(best_for_point["fc1"]),
                "delay": float(best_for_point["delay"]),
                "objective": float(best_for_point["objective"]),
                "event_threshold": float(best_for_point["event_threshold"]),
            }
        )

    return best_row, curve_rows


def plot_score_timeline(path: Path, dataset: str, scores: np.ndarray, labels: np.ndarray, point_threshold: float):
    path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(13, 4.8))
    x = np.arange(len(scores))
    ax.plot(x, scores, color="#1f4e79", linewidth=0.8, label="Point score")
    ax.axhline(point_threshold, color="#b22222", linestyle="--", linewidth=1.1, label=f"Point threshold={point_threshold:.4f}")

    in_event = False
    start = 0
    for idx, value in enumerate(labels):
        if value == 1 and not in_event:
            start = idx
            in_event = True
        elif value == 0 and in_event:
            ax.axvspan(start, idx - 1, color="#f4c542", alpha=0.25)
            in_event = False
    if in_event:
        ax.axvspan(start, len(labels) - 1, color="#f4c542", alpha=0.25)

    ax.set_title(f"{dataset} score-time curve with GT intervals")
    ax.set_xlabel("Time index")
    ax.set_ylabel("Score")
    ax.legend(loc="upper right")
    ax.grid(alpha=0.25, linestyle="--")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_threshold_curve(path: Path, dataset: str, curve: list[dict], chosen_threshold: float):
    path.parent.mkdir(parents=True, exist_ok=True)
    thresholds = np.array([row["point_threshold"] for row in curve], dtype=np.float64)
    f1_values = np.array([row["f1"] for row in curve], dtype=np.float64)
    fc1_values = np.array([row["fc1"] for row in curve], dtype=np.float64)
    objective_values = np.array([row["objective"] for row in curve], dtype=np.float64)

    fig, ax = plt.subplots(figsize=(7.8, 4.8))
    ax.plot(thresholds, f1_values, color="#1f4e79", linewidth=1.2, label="Unified F1")
    ax.plot(thresholds, fc1_values, color="#b22222", linewidth=1.2, label="Unified Fc1")
    ax.plot(thresholds, objective_values, color="#2a7f62", linewidth=1.2, label="Objective")
    ax.axvline(chosen_threshold, color="#704214", linestyle="--", linewidth=1.1, label=f"Chosen={chosen_threshold:.4f}")
    ax.set_title(f"{dataset} point-threshold vs F1/Fc1/objective")
    ax.set_xlabel("Point threshold")
    ax.set_ylabel("Value")
    ax.grid(alpha=0.25, linestyle="--")
    ax.legend(loc="best")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_event_overlay(path: Path, dataset: str, labels: np.ndarray, preds: np.ndarray):
    path.parent.mkdir(parents=True, exist_ok=True)
    true_events = points_to_events(labels)
    pred_events = points_to_events(preds)

    fig, ax = plt.subplots(figsize=(13, 2.6))
    ax.broken_barh([(start, end - start + 1) for start, end in true_events], (18, 8), facecolors="#f4c542")
    ax.broken_barh([(start, end - start + 1) for start, end in pred_events], (6, 8), facecolors="#1f4e79")
    ax.set_ylim(0, 30)
    ax.set_xlim(0, len(labels))
    ax.set_yticks([10, 22], labels=["Predicted", "Ground Truth"])
    ax.set_title(f"{dataset} predicted events vs true events")
    ax.grid(alpha=0.2, axis="x")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_merged_events(path: Path, dataset: str, raw_events: list[tuple[int, int]], final_events: list[tuple[int, int]], total_length: int):
    path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(13, 2.6))
    ax.broken_barh([(start, end - start + 1) for start, end in raw_events], (18, 8), facecolors="#d0d0d0")
    ax.broken_barh([(start, end - start + 1) for start, end in final_events], (6, 8), facecolors="#2a7f62")
    ax.set_ylim(0, 30)
    ax.set_xlim(0, total_length)
    ax.set_yticks([10, 22], labels=["Final events", "Raw events"])
    ax.set_title(f"{dataset} merged/final events vs raw events")
    ax.grid(alpha=0.2, axis="x")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_delay_distribution(path: Path, dataset: str, delays: list[float]):
    path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(7.2, 4.5))
    if delays:
        ax.hist(delays, bins=min(20, max(5, len(delays))), color="#1f4e79", alpha=0.85)
    else:
        ax.text(0.5, 0.5, "No detected events", ha="center", va="center", transform=ax.transAxes)
    ax.set_title(f"{dataset} delay distribution")
    ax.set_xlabel("Delay")
    ax.set_ylabel("Count")
    ax.grid(alpha=0.2, linestyle="--")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheets: list[tuple[str, list[dict]]]):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for idx, (sheet_name, rows) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet(title=sheet_name)
        ws.title = sheet_name
        header = list(rows[0].keys()) if rows else []
        ws.append(header)
        for row in rows:
            ws.append(list(row.values()))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 28)
        if header:
            end_col = ws.cell(row=1, column=ws.max_column).column_letter
            table = Table(displayName=f"{sheet_name}_table", ref=f"A1:{end_col}{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    args = parse_args()
    config = load_config(args.config)
    if args.datasets:
        requested = set(args.datasets)
        config["dataset_order"] = [dataset for dataset in config["dataset_order"] if dataset in requested]
        if not config["dataset_order"]:
            raise ValueError(f"No datasets from {args.datasets} matched config dataset_order.")
    analysis_dir = ROOT / config["analysis_dir"]
    results_csv = ROOT / config["results_csv"]
    comparison_csv = ROOT / config["comparison_csv"]
    results_xlsx = ROOT / config["results_xlsx"]
    official_csv = config.get("official_metrics_csv")
    official_rows = load_official_metrics(ROOT / official_csv if official_csv else None)
    calibration_cfg = config.get("threshold_calibration", {"enabled": False})

    detail_rows = []
    comparison_rows = []

    for dataset in config["dataset_order"]:
        print(f"processing {dataset}")
        score_root = resolve_score_root(dataset, config)
        loaded = load_dataset_inputs(dataset, score_root, trim_to_coverage=bool(config["use_alignment_trim"]))
        labels = loaded["labels"]
        window_point_scores = loaded["window_point_scores"]
        window_end_indices = loaded["window_end_indices"]

        baseline_cfg = config["baseline"]
        v1_cfg = config["event_aware_v1"]
        v2_cfg = config["event_aware_v2"]

        baseline_scores = build_point_scores(
            labels, window_point_scores, window_end_indices,
            aggregation=baseline_cfg["aggregation"],
            smoothing=baseline_cfg["smoothing"],
            smoothing_param=baseline_cfg["smoothing_param"],
        )

        v1_scores = build_point_scores(
            labels, window_point_scores, window_end_indices,
            aggregation=v1_cfg["aggregation"],
            smoothing=v1_cfg["smoothing"],
            smoothing_param=v1_cfg["smoothing_param"],
        )

        v2_scores = build_point_scores(
            labels, window_point_scores, window_end_indices,
            aggregation=v2_cfg["aggregation"],
            smoothing=v2_cfg["smoothing"],
            smoothing_param=v2_cfg["smoothing_param"],
        )
        baseline_views = build_threshold_views(
            dataset,
            labels,
            baseline_scores,
            calibration_cfg,
            baseline_cfg,
            score_root=score_root,
        )
        baseline_threshold = select_point_threshold(
            baseline_views["selection_scores"],
            baseline_views["selection_labels"],
            baseline_cfg,
            calibration_cfg,
        )
        baseline_preds_full = (baseline_scores >= baseline_threshold).astype(np.int64)
        baseline_eval_offset = baseline_views["evaluation_offset"]
        baseline_results = evaluate_predictions(
            labels[baseline_eval_offset:],
            baseline_preds_full[baseline_eval_offset:],
            baseline_scores[baseline_eval_offset:],
            threshold=baseline_threshold,
        )

        v1_views = build_threshold_views(
            dataset,
            labels,
            v1_scores,
            calibration_cfg,
            v1_cfg,
            score_root=score_root,
        )
        v1_threshold = select_point_threshold(
            v1_views["selection_scores"],
            v1_views["selection_labels"],
            v1_cfg,
            calibration_cfg,
        )
        v1_preds_full = (v1_scores >= v1_threshold).astype(np.int64)
        v1_eval_offset = v1_views["evaluation_offset"]
        v1_results = evaluate_predictions(
            labels[v1_eval_offset:],
            v1_preds_full[v1_eval_offset:],
            v1_scores[v1_eval_offset:],
            threshold=v1_threshold,
        )

        v2_views = build_threshold_views(
            dataset,
            labels,
            v2_scores,
            calibration_cfg,
            v2_cfg,
            score_root=score_root,
        )
        best_v2, curve = select_event_aware_v2_thresholds(
            v2_views["selection_scores"],
            v2_views["selection_labels"],
            v2_cfg,
            calibration_cfg,
        )
        v2_applied = apply_event_aware_thresholds(
            point_scores=v2_scores,
            total_length=len(labels),
            cfg=v2_cfg,
            point_threshold=best_v2["point_threshold"],
            event_threshold=best_v2["event_threshold"],
        )
        v2_preds_full = v2_applied["preds"]
        v2_processed = v2_applied["processed"]
        v2_eval_offset = v2_views["evaluation_offset"]
        v2_results = evaluate_predictions(
            labels[v2_eval_offset:],
            v2_preds_full[v2_eval_offset:],
            v2_scores[v2_eval_offset:],
            threshold=best_v2["point_threshold"],
        )

        plot_score_timeline(
            analysis_dir / f"{dataset}_score_time.png",
            dataset,
            v2_scores[v2_eval_offset:],
            labels[v2_eval_offset:],
            best_v2["point_threshold"],
        )
        plot_threshold_curve(analysis_dir / f"{dataset}_threshold_curve.png", dataset, curve, best_v2["point_threshold"])
        plot_event_overlay(
            analysis_dir / f"{dataset}_pred_vs_true_events.png",
            dataset,
            labels[v2_eval_offset:],
            v2_preds_full[v2_eval_offset:],
        )
        plot_merged_events(
            analysis_dir / f"{dataset}_merged_events_vs_raw_events.png",
            dataset,
            v2_processed["raw_events"],
            v2_processed["final_events"],
            len(labels),
        )
        plot_delay_distribution(
            analysis_dir / f"{dataset}_delay_distribution.png",
            dataset,
            event_detection_delays(labels[v2_eval_offset:], v2_preds_full[v2_eval_offset:]),
        )

        official = official_rows.get(dataset, default_official_metrics())
        detail_rows.append(
            {
                "Dataset": dataset,
                "Official PA-F1": format_value(official["PA-F1"]),
                "Official VUS-PR": format_value(official["VUS-PR"]),
                "Official VUS-ROC": format_value(official["VUS-ROC"]),
                "Score Source": str(score_root.relative_to(ROOT)),
                "Aligned Coverage Length": loaded["covered_length"],
                "Original Label Length": loaded["original_length"],
                "Threshold Calibration Mode": calibration_cfg.get("mode", "disabled"),
                "Threshold Selection Source": v2_views.get("selection_source", "unknown"),
                "Threshold Selection Length": int(v2_views["selection_length"]),
                "Evaluation Offset": int(v2_eval_offset),
                "Holdout Evaluation": str(bool(v2_views["evaluate_on_holdout_only"])),
                "Baseline Unified F1": format_value(baseline_results["f1"]),
                "Baseline Unified Fc1": format_value(baseline_results["fc1"]),
                "Baseline Unified Delay": format_value(baseline_results["delay"]),
                "V1 Unified F1": format_value(v1_results["f1"]),
                "V1 Unified Fc1": format_value(v1_results["fc1"]),
                "V1 Unified Delay": format_value(v1_results["delay"]),
                "V2 Unified F1": format_value(v2_results["f1"]),
                "V2 Unified Fc1": format_value(v2_results["fc1"]),
                "V2 Unified Delay": format_value(v2_results["delay"]),
                "V2 Objective": format_value(best_v2["objective"]),
                "V2 Point Threshold": format_value(best_v2["point_threshold"], digits=6),
                "V2 Event Threshold": format_value(best_v2["event_threshold"], digits=6),
                "V2 Aggregation": v2_cfg["aggregation"],
                "V2 Smoothing": v2_cfg["smoothing"],
                "V2 Smoothing Param": v2_cfg["smoothing_param"],
                "V2 Gap Merge": str(bool(v2_cfg["gap_merge"])),
                "V2 Gap Size": int(v2_cfg["gap_size"]),
                "V2 Min Event Length": int(v2_cfg["min_event_length"]),
                "V2 Event Score Pooling": v2_cfg["event_score_pooling"],
                "V2 Threshold Objective": v2_cfg["threshold_objective"],
                "V2 Threshold Lambda": float(v2_cfg["threshold_objective_lambda"]),
            }
        )
        comparison_rows.append(
            {
                "Dataset": dataset,
                "baseline_unified_f1": format_value(baseline_results["f1"]),
                "v1_unified_f1": format_value(v1_results["f1"]),
                "v2_unified_f1": format_value(v2_results["f1"]),
                "baseline_unified_fc1": format_value(baseline_results["fc1"]),
                "v1_unified_fc1": format_value(v1_results["fc1"]),
                "v2_unified_fc1": format_value(v2_results["fc1"]),
                "baseline_unified_delay": format_value(baseline_results["delay"]),
                "v1_unified_delay": format_value(v1_results["delay"]),
                "v2_unified_delay": format_value(v2_results["delay"]),
            }
        )
        print(f"finished {dataset}")

    write_csv(results_csv, list(detail_rows[0].keys()), detail_rows)
    write_csv(comparison_csv, list(comparison_rows[0].keys()), comparison_rows)
    write_xlsx(results_xlsx, [("event_aware_v2_results", detail_rows), ("comparison", comparison_rows)])
    print(results_csv)
    print(comparison_csv)
    print(results_xlsx)
    print(analysis_dir)


if __name__ == "__main__":
    main()
