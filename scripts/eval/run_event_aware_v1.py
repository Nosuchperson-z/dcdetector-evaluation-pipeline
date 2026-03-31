import argparse
import csv
import json
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from unified_evaluator import (
    aggregate_window_point_scores_to_points,
    apply_smoothing,
    evaluate_point_scores,
    points_to_events,
    threshold_candidates,
    threshold_scores,
)


ROOT = Path(__file__).resolve().parents[2]
PROCESSED_ROOT = ROOT / "data_processed"
TABLE_DIR = ROOT / "outputs" / "tables"
OFFICIAL_TABLE_PATH = TABLE_DIR / "official_raw_metrics.csv"


def parse_args():
    parser = argparse.ArgumentParser(description="Run Event-aware DCdetector v1 on saved score artifacts.")
    parser.add_argument("--config", default=str(ROOT / "configs" / "eval" / "event_aware_v1.json"))
    parser.add_argument("--use_event_aware", default=None, choices=["true", "false"])
    return parser.parse_args()


def load_config(path: str) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def load_official_metrics() -> dict[str, dict[str, str]]:
    with OFFICIAL_TABLE_PATH.open("r", encoding="utf-8", newline="") as fp:
        rows = list(csv.DictReader(fp))
    return {row["Dataset"]: row for row in rows}


def resolve_score_file(dataset: str, score_root: Path, filename: str) -> Path:
    dataset_dir = score_root / dataset
    direct = dataset_dir / filename
    if direct.exists():
        return direct
    matches = sorted(dataset_dir.glob(filename.replace(".npy", "*.npy")))
    if not matches:
        raise FileNotFoundError(f"Missing score file for {dataset}: {filename}")
    return matches[0]


def load_dataset_inputs(dataset: str, score_root: Path):
    labels = np.load(PROCESSED_ROOT / dataset / "label.npy").astype(np.int64)
    window_point_scores = np.load(resolve_score_file(dataset, score_root, "test_window_point_scores.npy")).astype(np.float32)
    window_end_indices = np.load(resolve_score_file(dataset, score_root, "test_window_end_indices.npy")).astype(np.int64).reshape(-1)
    effective_length = min(window_point_scores.shape[0], window_end_indices.shape[0])
    return labels, window_point_scores[:effective_length], window_end_indices[:effective_length]


def build_point_scores(
    labels: np.ndarray,
    window_point_scores: np.ndarray,
    window_end_indices: np.ndarray,
    aggregation: str,
    smoothing: str,
    smoothing_param: float,
) -> np.ndarray:
    point_scores = aggregate_window_point_scores_to_points(
        window_point_scores=window_point_scores,
        window_end_indices=window_end_indices,
        total_length=len(labels),
        method=aggregation,
    )
    return apply_smoothing(point_scores, smoothing, smoothing_param)


def threshold_curve(scores: np.ndarray, labels: np.ndarray, steps: int) -> list[dict]:
    curve = []
    for threshold in threshold_candidates(scores, steps=steps):
        results = evaluate_point_scores(scores, labels, float(threshold))
        curve.append(
            {
                "threshold": float(threshold),
                "f1": float(results["f1"]),
                "fc1": float(results["fc1"]),
                "delay": float(results["delay"]),
            }
        )
    return curve


def format_value(value, digits: int = 4):
    if isinstance(value, str):
        return value
    if not np.isfinite(value):
        return "inf"
    return f"{float(value):.{digits}f}"


def plot_score_timeline(path: Path, dataset: str, scores: np.ndarray, labels: np.ndarray, threshold: float):
    path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(13, 4.8))
    x = np.arange(len(scores))
    ax.plot(x, scores, color="#1f4e79", linewidth=0.8, label="Point score")
    ax.axhline(threshold, color="#b22222", linestyle="--", linewidth=1.1, label=f"Threshold={threshold:.4f}")

    in_event = False
    start = 0
    for idx, value in enumerate(labels):
        if value == 1 and not in_event:
            start = idx
            in_event = True
        elif value == 0 and in_event:
            ax.axvspan(start, idx - 1, color="#f4c542", alpha=0.25)
            in_event = False
    if in_event:
        ax.axvspan(start, len(labels) - 1, color="#f4c542", alpha=0.25)

    ax.set_title(f"{dataset} score-time curve with GT intervals")
    ax.set_xlabel("Time index")
    ax.set_ylabel("Score")
    ax.legend(loc="upper right")
    ax.grid(alpha=0.25, linestyle="--")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_threshold_curve(path: Path, dataset: str, curve: list[dict], chosen_threshold: float):
    path.parent.mkdir(parents=True, exist_ok=True)
    thresholds = np.array([row["threshold"] for row in curve], dtype=np.float64)
    f1_values = np.array([row["f1"] for row in curve], dtype=np.float64)
    fc1_values = np.array([row["fc1"] for row in curve], dtype=np.float64)

    fig, ax = plt.subplots(figsize=(7.5, 4.8))
    ax.plot(thresholds, f1_values, color="#1f4e79", linewidth=1.2, label="Unified F1")
    ax.plot(thresholds, fc1_values, color="#b22222", linewidth=1.2, label="Unified Fc1")
    ax.axvline(chosen_threshold, color="#2a7f62", linestyle="--", linewidth=1.1, label=f"Chosen={chosen_threshold:.4f}")
    ax.set_title(f"{dataset} threshold vs F1/Fc1")
    ax.set_xlabel("Threshold")
    ax.set_ylabel("Metric value")
    ax.grid(alpha=0.25, linestyle="--")
    ax.legend(loc="best")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_event_overlay(path: Path, dataset: str, labels: np.ndarray, preds: np.ndarray):
    path.parent.mkdir(parents=True, exist_ok=True)
    true_events = points_to_events(labels)
    pred_events = points_to_events(preds)

    fig, ax = plt.subplots(figsize=(13, 2.6))
    ax.broken_barh([(start, end - start + 1) for start, end in true_events], (18, 8), facecolors="#f4c542")
    ax.broken_barh([(start, end - start + 1) for start, end in pred_events], (6, 8), facecolors="#1f4e79")
    ax.set_ylim(0, 30)
    ax.set_xlim(0, len(labels))
    ax.set_yticks([10, 22], labels=["Predicted", "Ground Truth"])
    ax.set_title(f"{dataset} predicted events vs true events")
    ax.grid(alpha=0.2, axis="x")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheets: list[tuple[str, list[dict]]]):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for idx, (sheet_name, rows) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet(title=sheet_name)
        ws.title = sheet_name
        header = list(rows[0].keys()) if rows else []
        ws.append(header)
        for row in rows:
            ws.append(list(row.values()))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 26)
        if header:
            end_col = ws.cell(row=1, column=ws.max_column).column_letter
            table = Table(displayName=f"{sheet_name}_table", ref=f"A1:{end_col}{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    args = parse_args()
    config = load_config(args.config)
    if args.use_event_aware is not None:
        config["use_event_aware"] = args.use_event_aware == "true"

    score_root = ROOT / config["score_root"]
    analysis_dir = ROOT / config["analysis_dir"]
    results_csv = ROOT / config["results_csv"]
    comparison_csv = ROOT / config["comparison_csv"]
    results_xlsx = ROOT / config["results_xlsx"]

    baseline_cfg = config["baseline"]
    event_cfg = config["event_aware"] if config["use_event_aware"] else config["baseline"]
    official_rows = load_official_metrics()

    detail_rows = []
    comparison_rows = []

    for dataset in config["dataset_order"]:
        labels, window_point_scores, window_end_indices = load_dataset_inputs(dataset, score_root)

        baseline_scores = build_point_scores(
            labels,
            window_point_scores,
            window_end_indices,
            aggregation=baseline_cfg["aggregation"],
            smoothing=baseline_cfg["smoothing"],
            smoothing_param=baseline_cfg["smoothing_param"],
        )
        baseline_threshold = threshold_scores(
            baseline_scores,
            baseline_cfg["threshold_method"],
            labels=labels,
            param=baseline_cfg["threshold_param"],
        )
        baseline_results = evaluate_point_scores(baseline_scores, labels, baseline_threshold)

        event_scores = build_point_scores(
            labels,
            window_point_scores,
            window_end_indices,
            aggregation=event_cfg["aggregation"],
            smoothing=event_cfg["smoothing"],
            smoothing_param=event_cfg["smoothing_param"],
        )
        event_threshold = threshold_scores(
            event_scores,
            event_cfg["threshold_method"],
            labels=labels,
            param=event_cfg["threshold_param"],
        )
        event_results = evaluate_point_scores(event_scores, labels, event_threshold)

        curve_steps = int(event_cfg["threshold_param"]) if event_cfg["threshold_method"] == "best_fc1" and event_cfg["threshold_param"] else 401
        curve = threshold_curve(event_scores, labels, steps=curve_steps)
        event_preds = (event_scores >= event_threshold).astype(np.int64)

        plot_score_timeline(analysis_dir / f"{dataset}_score_time.png", dataset, event_scores, labels, event_threshold)
        plot_threshold_curve(analysis_dir / f"{dataset}_threshold_curve.png", dataset, curve, event_threshold)
        plot_event_overlay(analysis_dir / f"{dataset}_pred_vs_true_events.png", dataset, labels, event_preds)

        official = official_rows[dataset]
        detail_rows.append(
            {
                "Dataset": dataset,
                "Official PA-F1": format_value(float(official["PA-F1"])),
                "Official VUS-PR": format_value(float(official["VUS-PR"])),
                "Official VUS-ROC": format_value(float(official["VUS-ROC"])),
                "Baseline Unified F1": format_value(baseline_results["f1"]),
                "Baseline Unified Fc1": format_value(baseline_results["fc1"]),
                "Baseline Unified Delay": format_value(baseline_results["delay"]),
                "Event-aware Unified F1": format_value(event_results["f1"]),
                "Event-aware Unified Fc1": format_value(event_results["fc1"]),
                "Event-aware Unified Delay": format_value(event_results["delay"]),
                "Event-aware Threshold": format_value(event_results["threshold"], digits=6),
                "Aggregation": event_cfg["aggregation"],
                "Smoothing": event_cfg["smoothing"],
                "Smoothing Param": event_cfg["smoothing_param"],
                "Threshold Method": event_cfg["threshold_method"],
                "Use Event-aware": str(bool(config["use_event_aware"])),
            }
        )

        comparison_rows.append(
            {
                "Dataset": dataset,
                "baseline_unified_f1": format_value(baseline_results["f1"]),
                "event_aware_unified_f1": format_value(event_results["f1"]),
                "baseline_unified_fc1": format_value(baseline_results["fc1"]),
                "event_aware_unified_fc1": format_value(event_results["fc1"]),
                "baseline_unified_delay": format_value(baseline_results["delay"]),
                "event_aware_unified_delay": format_value(event_results["delay"]),
            }
        )

    write_csv(results_csv, list(detail_rows[0].keys()), detail_rows)
    write_csv(comparison_csv, list(comparison_rows[0].keys()), comparison_rows)
    write_xlsx(results_xlsx, [("event_aware_results", detail_rows), ("comparison", comparison_rows)])
    print(results_csv)
    print(comparison_csv)
    print(results_xlsx)
    print(analysis_dir)


if __name__ == "__main__":
    main()
